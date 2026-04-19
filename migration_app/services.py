from __future__ import annotations

import csv
import io
import json
import re
import uuid
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from finance.models import Invoice, OpeningBalance, Payment
from masters.models import Party
from migration_app.models import MigrationBatch, MigrationMappingProfile, MigrationRowError
from production.models import Challan


ZERO = Decimal("0.00")
WORKBOOK_LEDGER_SHEET = "Ledger Report"
WORKBOOK_PRODUCTION_SHEET = "Daily Production Matrix"
WORKBOOK_PARTY_ALIASES = {
    "AGGARWAL": "AGARWAL IRON INDUSTRIES",
    "DERBY": "DERBY INDUSTRIES.",
    "FITWELL": "FITWELL FASTNERS & FITTING INDUSTRIES",
    "GINTER": "GINTER FORGING P. LTD.",
    "KV": "K.V.ENTERPRISES",
    "KIC": "KIC",
    "KOMAL": "KOMAL INDUSTRIAL",
    "LOYAL": "LOYAL ENTERPRISES",
    "MALHOTRA": "MALHOTRA INDL. CORPN.",
    "NAMDHARI": "NAMDHARI INDUSTRIAL TRADERS PVT.LTD.",
    "PRECISION": "PRECISION AUTO FASTNERS",
    "PREMCO": "PREMCO INDIA EXIM PVT. LTD.",
    "RADHA": "RADHA INDUSTRIES",
    "UDHERA": "UDEHRA FASTNERS",
    "KAPOOR SONS": "KAPOOR SONS",
    "KIRTAN": "KIRTAN",
}


DEFAULT_MAPPINGS = {
    MigrationBatch.ImportType.PARTIES: {
        "name": "name",
        "contact_person": "contact_person",
        "phone": "phone",
        "gst_number": "gst_number",
        "address": "address",
        "is_active": "is_active",
    },
    MigrationBatch.ImportType.OPENING_BALANCES: {
        "party": "party",
        "effective_date": "effective_date",
        "balance_type": "balance_type",
        "amount": "amount",
        "remarks": "remarks",
    },
    MigrationBatch.ImportType.INVOICES: {
        "party": "party",
        "invoice_number": "invoice_number",
        "invoice_date": "invoice_date",
        "amount": "amount",
        "remarks": "remarks",
    },
    MigrationBatch.ImportType.PAYMENTS: {
        "party": "party",
        "payment_date": "payment_date",
        "amount": "amount",
        "mode": "mode",
        "reference_number": "reference_number",
        "remarks": "remarks",
    },
    MigrationBatch.ImportType.CHALLANS: {
        "party": "party",
        "challan_number": "challan_number",
        "challan_date": "challan_date",
        "job_description": "job_description",
        "job_type": "job_type",
        "direction": "direction",
        "weight_kg": "weight_kg",
        "amount": "amount",
        "remarks": "remarks",
    },
}


@dataclass
class PreviewRow:
    row_number: int
    sheet_name: str
    normalized: dict[str, Any]
    duplicate_warning: str = ""


@dataclass
class PreviewError:
    row_number: int
    sheet_name: str
    raw_payload: dict[str, Any]
    error_message: str


@dataclass
class ImportPreview:
    token: str
    import_type: str
    source_file_name: str
    file_type: str
    valid_rows: list[PreviewRow]
    errors: list[PreviewError]

    @property
    def projected_totals(self) -> dict[str, Any]:
        parties = sorted(
            {
                row.normalized.get("party", "") or row.normalized.get("name", "")
                for row in self.valid_rows
                if row.normalized.get("party") or row.normalized.get("name")
            }
        )
        total_amount = sum(Decimal(str(row.normalized.get("amount", "0") or "0")) for row in self.valid_rows if "amount" in row.normalized)
        record_counts: dict[str, int] = {}
        for row in self.valid_rows:
            record_type = str(row.normalized.get("record_type", self.import_type))
            record_counts[record_type] = record_counts.get(record_type, 0) + 1
        return {
            "valid_row_count": len(self.valid_rows),
            "error_count": len(self.errors),
            "affected_parties": parties,
            "total_amount": str(total_amount),
            "record_counts": record_counts,
        }


def serialize_preview(preview: ImportPreview) -> dict[str, Any]:
    def normalize_value(value: Any) -> Any:
        if isinstance(value, (date, datetime, Decimal)):
            return str(value)
        return value

    def serialize_row(row: PreviewRow) -> dict[str, Any]:
        payload = asdict(row)
        payload["normalized"] = {key: normalize_value(value) for key, value in row.normalized.items()}
        return payload

    return {
        "token": preview.token,
        "import_type": preview.import_type,
        "source_file_name": preview.source_file_name,
        "file_type": preview.file_type,
        "valid_rows": [serialize_row(row) for row in preview.valid_rows],
        "errors": [asdict(error) for error in preview.errors],
    }


def deserialize_preview(payload: dict[str, Any]) -> ImportPreview:
    return ImportPreview(
        token=payload["token"],
        import_type=payload["import_type"],
        source_file_name=payload["source_file_name"],
        file_type=payload["file_type"],
        valid_rows=[PreviewRow(**row) for row in payload.get("valid_rows", [])],
        errors=[PreviewError(**row) for row in payload.get("errors", [])],
    )


def mapping_for(import_type: str, profile: MigrationMappingProfile | None) -> dict[str, str]:
    if profile and profile.import_type == import_type and profile.column_mapping_config:
        return profile.column_mapping_config
    return DEFAULT_MAPPINGS[import_type]


def normalize_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date '{value}'")


def normalize_decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        raise ValueError(f"Invalid amount '{value}'")


def normalize_boolean(value: Any) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() not in {"0", "false", "no", ""}


def get_upload_rows(upload) -> tuple[list[tuple[str, int, dict[str, Any]]], str]:
    suffix = Path(upload.name).suffix.lower()
    if suffix == ".csv":
        content = upload.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(content))
        rows = [("CSV", idx, row) for idx, row in enumerate(reader, start=2)]
        return rows, MigrationBatch.FileType.CSV

    workbook = load_workbook(upload, data_only=True)
    parsed_rows: list[tuple[str, int, dict[str, Any]]] = []
    for worksheet in workbook.worksheets:
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            parsed_rows.append((worksheet.title, row_index, {headers[i]: value for i, value in enumerate(row)}))
    return parsed_rows, MigrationBatch.FileType.XLSX


def workbook_party_name(name: Any) -> str:
    canonical = str(name or "").strip()
    if not canonical:
        raise ValueError("Party is required")
    alias_key = re.sub(r"\s+", " ", canonical.upper())
    return WORKBOOK_PARTY_ALIASES.get(alias_key, canonical.upper())


def payment_mode_from_particulars(particulars: str) -> str:
    upper = particulars.upper()
    if "CREDIT" in upper or "R/D" in upper:
        return Payment.Mode.ADJUSTMENT
    if "CH.NO" in upper or "CHEQUE" in upper:
        return Payment.Mode.CHEQUE
    if "CASH" in upper:
        return Payment.Mode.CASH
    return Payment.Mode.BANK


def invoice_number_from_particulars(particulars: str) -> str:
    match = re.search(r"BILL NO\.?\s*(.+)$", particulars, re.IGNORECASE)
    return match.group(1).strip() if match else ""


def payment_reference_from_particulars(particulars: str) -> str:
    match = re.search(r"CH\.?NO\.?\s*([A-Z0-9/-]+)", particulars, re.IGNORECASE)
    if match:
        reference = match.group(1).strip().upper()
        if reference in {"TRF", "NEFT", "RTGS", "R/D"}:
            return ""
        if not any(character.isdigit() for character in reference):
            return ""
        return reference
    return ""


def invoice_identifier_needs_suffix(party_name: str, invoice_number: str, invoice_date: date, amount: Decimal) -> bool:
    if not invoice_number:
        return False
    return Invoice.objects.filter(party__name=party_name, invoice_number=invoice_number).exclude(
        invoice_date=invoice_date,
        amount=amount,
    ).exists()


def payment_reference_needs_suffix(party_name: str, reference_number: str, payment_date: date, amount: Decimal) -> bool:
    if not reference_number:
        return False
    return Payment.objects.filter(party__name=party_name, reference_number=reference_number).exclude(
        payment_date=payment_date,
        amount=amount,
    ).exists()


def workbook_record_import_type(normalized: dict[str, Any]) -> str:
    record_type = normalized.get("record_type")
    if record_type == "party":
        return MigrationBatch.ImportType.PARTIES
    if record_type == "opening_balance":
        return MigrationBatch.ImportType.OPENING_BALANCES
    if record_type == "invoice":
        return MigrationBatch.ImportType.INVOICES
    if record_type == "payment":
        return MigrationBatch.ImportType.PAYMENTS
    if record_type == "challan":
        return MigrationBatch.ImportType.CHALLANS
    raise ValueError(f"Unsupported workbook record type '{record_type}'")


def detect_duplicate(import_type: str, normalized: dict[str, Any]) -> str:
    if import_type == MigrationBatch.ImportType.PARTIES:
        if Party.objects.filter(name=normalized["name"]).exists():
            return "Existing party name"
    elif import_type == MigrationBatch.ImportType.OPENING_BALANCES:
        if OpeningBalance.objects.filter(
            party=normalized["party_obj"],
            effective_date=normalized["effective_date"],
            balance_type=normalized["balance_type"],
            amount=normalized["amount"],
        ).exists():
            return "Matching opening balance already exists"
    elif import_type == MigrationBatch.ImportType.INVOICES:
        filters = {
            "party": normalized["party_obj"],
            "invoice_date": normalized["invoice_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("invoice_number"):
            filters["invoice_number"] = normalized["invoice_number"]
        if Invoice.objects.filter(**filters).exists():
            return "Matching invoice already exists"
    elif import_type == MigrationBatch.ImportType.PAYMENTS:
        filters = {
            "party": normalized["party_obj"],
            "payment_date": normalized["payment_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("reference_number"):
            filters["reference_number"] = normalized["reference_number"]
        if Payment.objects.filter(**filters).exists():
            return "Matching payment already exists"
    elif import_type == MigrationBatch.ImportType.CHALLANS:
        filters = {
            "party": normalized["party_obj"],
            "challan_date": normalized["challan_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("challan_number"):
            filters["challan_number"] = normalized["challan_number"]
        if Challan.objects.filter(**filters).exists():
            return "Matching challan already exists"
    return ""


def detect_workbook_duplicate(normalized: dict[str, Any]) -> str:
    record_type = normalized.get("record_type")
    if record_type == "party":
        if Party.objects.filter(name=normalized["name"]).exists():
            return "Existing party name"
        return ""

    if record_type == "opening_balance":
        if OpeningBalance.objects.filter(
            party__name=normalized["party"],
            effective_date=normalized["effective_date"],
            balance_type=normalized["balance_type"],
            amount=normalized["amount"],
        ).exists():
            return "Matching opening balance already exists"
        return ""

    if record_type == "invoice":
        filters = {
            "party__name": normalized["party"],
            "invoice_date": normalized["invoice_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("invoice_number"):
            filters["invoice_number"] = normalized["invoice_number"]
        if Invoice.objects.filter(**filters).exists():
            return "Matching invoice already exists"
        return ""

    if record_type == "payment":
        filters = {
            "party__name": normalized["party"],
            "payment_date": normalized["payment_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("reference_number"):
            filters["reference_number"] = normalized["reference_number"]
        if Payment.objects.filter(**filters).exists():
            return "Matching payment already exists"
        return ""

    if record_type == "challan":
        filters = {
            "party__name": normalized["party"],
            "challan_date": normalized["challan_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("challan_number"):
            filters["challan_number"] = normalized["challan_number"]
        if Challan.objects.filter(**filters).exists():
            return "Matching challan already exists"
        return ""

    raise ValueError(f"Unsupported workbook record type '{record_type}'")


def normalize_row(import_type: str, raw_payload: dict[str, Any], profile: MigrationMappingProfile | None) -> dict[str, Any]:
    mapping = mapping_for(import_type, profile)
    normalized: dict[str, Any] = {}
    for source_column, target_field in mapping.items():
        normalized[target_field] = raw_payload.get(source_column)

    if import_type == MigrationBatch.ImportType.PARTIES:
        name = str(normalized.get("name", "")).strip()
        if not name:
            raise ValueError("Party name is required")
        normalized["name"] = name
        normalized["is_active"] = normalize_boolean(normalized.get("is_active", True))
        return normalized

    party_name = str(normalized.get("party", "")).strip()
    if not party_name:
        raise ValueError("Party is required")
    try:
        normalized["party_obj"] = Party.objects.get(name=party_name)
    except Party.DoesNotExist as exc:
        raise ValueError(f"Unknown party '{party_name}'") from exc

    if import_type == MigrationBatch.ImportType.OPENING_BALANCES:
        normalized["effective_date"] = normalize_date(normalized.get("effective_date"))
        normalized["amount"] = normalize_decimal(normalized.get("amount"))
        balance_type = str(normalized.get("balance_type", "DR")).strip().upper()
        if balance_type not in {OpeningBalance.BalanceType.DEBIT, OpeningBalance.BalanceType.CREDIT}:
            raise ValueError("Balance type must be DR or CR")
        normalized["balance_type"] = balance_type
    elif import_type == MigrationBatch.ImportType.INVOICES:
        normalized["invoice_number"] = str(normalized.get("invoice_number", "") or "").strip()
        normalized["invoice_date"] = normalize_date(normalized.get("invoice_date"))
        normalized["amount"] = normalize_decimal(normalized.get("amount"))
    elif import_type == MigrationBatch.ImportType.PAYMENTS:
        normalized["payment_date"] = normalize_date(normalized.get("payment_date"))
        normalized["amount"] = normalize_decimal(normalized.get("amount"))
        normalized["mode"] = str(normalized.get("mode", Payment.Mode.BANK)).strip().lower() or Payment.Mode.BANK
        if normalized["mode"] not in Payment.Mode.values:
            raise ValueError(f"Unsupported payment mode '{normalized['mode']}'")
        normalized["reference_number"] = str(normalized.get("reference_number", "") or "").strip()
    elif import_type == MigrationBatch.ImportType.CHALLANS:
        normalized["challan_number"] = str(normalized.get("challan_number", "") or "").strip()
        normalized["challan_date"] = normalize_date(normalized.get("challan_date"))
        normalized["job_description"] = str(normalized.get("job_description", "") or "").strip()
        if not normalized["job_description"]:
            raise ValueError("Job description is required")
        normalized["job_type"] = str(normalized.get("job_type", "") or "").strip()
        normalized["direction"] = str(normalized.get("direction", Challan.Direction.INBOUND)).strip().upper() or Challan.Direction.INBOUND
        if normalized["direction"] not in Challan.Direction.values:
            raise ValueError("Direction must be IN or OUT")
        normalized["weight_kg"] = normalize_decimal(normalized.get("weight_kg", 0))
        normalized["amount"] = normalize_decimal(normalized.get("amount", 0))
    return normalized


def build_workbook_preview(upload) -> ImportPreview:
    suffix = Path(upload.name).suffix.lower()
    if suffix != ".xlsx":
        raise ValueError("Production dashboard workbook import requires an XLSX file.")

    workbook = load_workbook(upload, data_only=True)
    if WORKBOOK_LEDGER_SHEET not in workbook.sheetnames or WORKBOOK_PRODUCTION_SHEET not in workbook.sheetnames:
        raise ValueError("Workbook must contain both 'Ledger Report' and 'Daily Production Matrix' sheets.")

    valid_rows: list[PreviewRow] = []
    errors: list[PreviewError] = []
    seen_signatures: set[tuple[Any, ...]] = set()
    seen_invoice_numbers: set[tuple[str, str]] = set()
    seen_payment_references: set[tuple[str, str]] = set()

    ledger_sheet = workbook[WORKBOOK_LEDGER_SHEET]
    parties: set[str] = set()
    for row in ledger_sheet.iter_rows(min_row=2, values_only=True):
        party_name = row[2]
        if party_name:
            parties.add(workbook_party_name(party_name))

    production_sheet = workbook[WORKBOOK_PRODUCTION_SHEET]
    production_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in production_sheet[1]]
    for row_index, row in enumerate(production_sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_payload = {production_headers[i]: value for i, value in enumerate(row)}
        if not any(value is not None and value != "" for value in raw_payload.values()):
            continue
        try:
            parties.add(workbook_party_name(raw_payload.get("Party Name")))
        except ValueError as exc:
            errors.append(
                PreviewError(
                    row_number=row_index,
                    sheet_name=WORKBOOK_PRODUCTION_SHEET,
                    raw_payload={key: ("" if value is None else str(value)) for key, value in raw_payload.items()},
                    error_message=str(exc),
                )
            )

    for party_name in sorted(parties):
        normalized = {"record_type": "party", "name": party_name, "is_active": True}
        duplicate_warning = detect_duplicate(MigrationBatch.ImportType.PARTIES, normalized)
        valid_rows.append(
            PreviewRow(
                row_number=1,
                sheet_name="Party Master",
                normalized=normalized,
                duplicate_warning=duplicate_warning,
            )
        )

    ledger_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ledger_sheet[1]]
    for row_index, row in enumerate(ledger_sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_payload = {ledger_headers[i]: value for i, value in enumerate(row)}
        if not any(value is not None and value != "" for value in raw_payload.values()):
            continue
        try:
            entry_date = normalize_date(raw_payload.get("Date"))
            party_name = workbook_party_name(raw_payload.get("Party Name"))
            particulars = str(raw_payload.get("Particulars", "") or "").strip() or "Ledger import"
            credit_amount = normalize_decimal(raw_payload.get("Credit", 0) or 0)
            debit_amount = normalize_decimal(raw_payload.get("Debit", 0) or 0)
            uniqueness_warning = ""

            if particulars.upper().startswith("OPENING BALANCE"):
                if debit_amount == ZERO and credit_amount == ZERO:
                    continue
                normalized = {
                    "record_type": "opening_balance",
                    "party": party_name,
                    "effective_date": entry_date,
                    "balance_type": OpeningBalance.BalanceType.DEBIT if debit_amount >= credit_amount else OpeningBalance.BalanceType.CREDIT,
                    "amount": debit_amount if debit_amount >= credit_amount else credit_amount,
                    "remarks": particulars,
                }
            elif debit_amount > ZERO:
                invoice_number = invoice_number_from_particulars(particulars)
                uniqueness_warning = ""
                if invoice_number and (
                    (party_name, invoice_number) in seen_invoice_numbers
                    or invoice_identifier_needs_suffix(party_name, invoice_number, entry_date, debit_amount)
                ):
                    invoice_number = f"{invoice_number}__{entry_date.isoformat()}"
                    uniqueness_warning = "Source invoice number repeated; stored with date suffix for uniqueness"
                if invoice_number:
                    seen_invoice_numbers.add((party_name, invoice_number))
                normalized = {
                    "record_type": "invoice",
                    "party": party_name,
                    "invoice_number": invoice_number,
                    "invoice_date": entry_date,
                    "amount": debit_amount,
                    "remarks": particulars,
                }
            elif credit_amount > ZERO:
                reference_number = payment_reference_from_particulars(particulars)
                uniqueness_warning = ""
                if reference_number and (
                    (party_name, reference_number) in seen_payment_references
                    or payment_reference_needs_suffix(party_name, reference_number, entry_date, credit_amount)
                ):
                    reference_number = f"{reference_number}__{entry_date.isoformat()}"
                    uniqueness_warning = "Source payment reference repeated; stored with date suffix for uniqueness"
                if reference_number:
                    seen_payment_references.add((party_name, reference_number))
                normalized = {
                    "record_type": "payment",
                    "party": party_name,
                    "payment_date": entry_date,
                    "amount": credit_amount,
                    "mode": payment_mode_from_particulars(particulars),
                    "reference_number": reference_number,
                    "remarks": particulars,
                }
            else:
                continue

            signature = tuple(sorted((key, str(value)) for key, value in normalized.items()))
            duplicate_warning = "Duplicate row in workbook" if signature in seen_signatures else detect_workbook_duplicate(normalized)
            if uniqueness_warning:
                duplicate_warning = f"{duplicate_warning} | {uniqueness_warning}".strip(" |")
            seen_signatures.add(signature)
            valid_rows.append(
                PreviewRow(
                    row_number=row_index,
                    sheet_name=WORKBOOK_LEDGER_SHEET,
                    normalized=normalized,
                    duplicate_warning=duplicate_warning,
                )
            )
        except ValueError as exc:
            errors.append(
                PreviewError(
                    row_number=row_index,
                    sheet_name=WORKBOOK_LEDGER_SHEET,
                    raw_payload={key: ("" if value is None else str(value)) for key, value in raw_payload.items()},
                    error_message=str(exc),
                )
            )

    for row_index, row in enumerate(production_sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_payload = {production_headers[i]: value for i, value in enumerate(row)}
        if not any(value is not None and value != "" for value in raw_payload.values()):
            continue
        try:
            challan_date = normalize_date(raw_payload.get("Date"))
            party_name = workbook_party_name(raw_payload.get("Party Name"))
            job_description = str(raw_payload.get("Material", "") or "").strip()
            if not job_description:
                raise ValueError("Job description is required")
            weight_kg = normalize_decimal(raw_payload.get("Kg(s)", 0) or 0)
            amount = normalize_decimal(raw_payload.get("Production Output (₹)", 0) or 0)
            serial = str(raw_payload.get("S. No.", "") or "").strip()
            remarks_parts = [
                f"Source row {serial}" if serial else "",
                f"Shift: {raw_payload.get('Shift')}" if raw_payload.get("Shift") else "",
                f"Rate: {raw_payload.get('Rate')}" if raw_payload.get("Rate") not in (None, "") else "",
                f"Workers: {raw_payload.get('No. of Workers')}" if raw_payload.get("No. of Workers") not in (None, "") else "",
                f"Hours: {raw_payload.get('Working Hours')}" if raw_payload.get("Working Hours") not in (None, "") else "",
                f"Remarks: {raw_payload.get('Remarks')}" if raw_payload.get("Remarks") else "",
            ]
            normalized = {
                "record_type": "challan",
                "party": party_name,
                "challan_number": "",
                "challan_date": challan_date,
                "job_description": job_description,
                "job_type": "Production dashboard import",
                "direction": Challan.Direction.OUTBOUND,
                "weight_kg": weight_kg,
                "amount": amount,
                "remarks": " | ".join(part for part in remarks_parts if part),
            }
            signature = tuple(sorted((key, str(value)) for key, value in normalized.items()))
            duplicate_warning = "Duplicate row in workbook" if signature in seen_signatures else detect_workbook_duplicate(normalized)
            seen_signatures.add(signature)
            valid_rows.append(
                PreviewRow(
                    row_number=row_index,
                    sheet_name=WORKBOOK_PRODUCTION_SHEET,
                    normalized=normalized,
                    duplicate_warning=duplicate_warning,
                )
            )
        except ValueError as exc:
            errors.append(
                PreviewError(
                    row_number=row_index,
                    sheet_name=WORKBOOK_PRODUCTION_SHEET,
                    raw_payload={key: ("" if value is None else str(value)) for key, value in raw_payload.items()},
                    error_message=str(exc),
                )
            )

    return ImportPreview(
        token=uuid.uuid4().hex,
        import_type=MigrationBatch.ImportType.PRODUCTION_DASHBOARD_WORKBOOK,
        source_file_name=upload.name,
        file_type=MigrationBatch.FileType.XLSX,
        valid_rows=valid_rows,
        errors=errors,
    )


def build_preview(upload, import_type: str, profile: MigrationMappingProfile | None = None) -> ImportPreview:
    if import_type == MigrationBatch.ImportType.PRODUCTION_DASHBOARD_WORKBOOK:
        return build_workbook_preview(upload)

    rows, file_type = get_upload_rows(upload)
    valid_rows: list[PreviewRow] = []
    errors: list[PreviewError] = []
    for sheet_name, row_number, raw_payload in rows:
        try:
            normalized = normalize_row(import_type, raw_payload, profile)
            duplicate_warning = detect_duplicate(import_type, normalized)
            for transient_key in ("party_obj",):
                normalized.pop(transient_key, None)
            valid_rows.append(
                PreviewRow(
                    row_number=row_number,
                    sheet_name=sheet_name,
                    normalized=normalized,
                    duplicate_warning=duplicate_warning,
                )
            )
        except ValueError as exc:
            errors.append(
                PreviewError(
                    row_number=row_number,
                    sheet_name=sheet_name,
                    raw_payload={key: ("" if value is None else str(value)) for key, value in raw_payload.items()},
                    error_message=str(exc),
                )
            )

    return ImportPreview(
        token=uuid.uuid4().hex,
        import_type=import_type,
        source_file_name=upload.name,
        file_type=file_type,
        valid_rows=valid_rows,
        errors=errors,
    )


def _persist_row(batch: MigrationBatch, import_type: str, row_number: int, normalized: dict[str, Any]) -> None:
    if import_type == MigrationBatch.ImportType.PARTIES:
        Party.objects.update_or_create(
            name=normalized["name"],
            defaults={
                "contact_person": normalized.get("contact_person", "") or "",
                "phone": normalized.get("phone", "") or "",
                "gst_number": normalized.get("gst_number", "") or "",
                "address": normalized.get("address", "") or "",
                "is_active": normalized.get("is_active", True),
                "source_batch": batch,
                "source_row_number": row_number,
            },
        )
        return

    normalized = normalize_row(import_type, normalized, profile=None)
    common_fields = {"source_batch": batch, "source_row_number": row_number}
    if import_type == MigrationBatch.ImportType.OPENING_BALANCES:
        OpeningBalance.objects.get_or_create(
            party=normalized["party_obj"],
            effective_date=normalized["effective_date"],
            balance_type=normalized["balance_type"],
            amount=normalized["amount"],
            defaults={"remarks": normalized.get("remarks", "") or "", **common_fields},
        )
    elif import_type == MigrationBatch.ImportType.INVOICES:
        lookup = {
            "party": normalized["party_obj"],
            "invoice_date": normalized["invoice_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("invoice_number"):
            lookup["invoice_number"] = normalized["invoice_number"]
        Invoice.objects.get_or_create(
            **lookup,
            defaults={"remarks": normalized.get("remarks", "") or "", **common_fields},
        )
    elif import_type == MigrationBatch.ImportType.PAYMENTS:
        lookup = {
            "party": normalized["party_obj"],
            "payment_date": normalized["payment_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("reference_number"):
            lookup["reference_number"] = normalized["reference_number"]
        Payment.objects.get_or_create(
            **lookup,
            defaults={"remarks": normalized.get("remarks", "") or "", "mode": normalized["mode"], **common_fields},
        )
    elif import_type == MigrationBatch.ImportType.CHALLANS:
        lookup = {
            "party": normalized["party_obj"],
            "challan_date": normalized["challan_date"],
            "amount": normalized["amount"],
        }
        if normalized.get("challan_number"):
            lookup["challan_number"] = normalized["challan_number"]
        Challan.objects.get_or_create(
            **lookup,
            defaults={
                "job_description": normalized["job_description"],
                "job_type": normalized.get("job_type", "") or "",
                "direction": normalized["direction"],
                "weight_kg": normalized["weight_kg"],
                "remarks": normalized.get("remarks", "") or "",
                **common_fields,
            },
        )


def _persist_workbook_row(batch: MigrationBatch, row_number: int, normalized: dict[str, Any]) -> None:
    _persist_row(batch, workbook_record_import_type(normalized), row_number, {key: value for key, value in normalized.items() if key != "record_type"})


@transaction.atomic
def commit_preview(preview: ImportPreview, upload_bytes: bytes, user) -> MigrationBatch:
    batch = MigrationBatch.objects.create(
        source_file_name=preview.source_file_name,
        import_type=preview.import_type,
        file_type=preview.file_type,
        status=MigrationBatch.Status.IMPORTED,
        row_count=len(preview.valid_rows) + len(preview.errors),
        success_count=len(preview.valid_rows),
        error_count=len(preview.errors),
        imported_at=timezone.now(),
        triggered_by=user if user.is_authenticated else None,
    )
    batch.upload.save(preview.source_file_name, ContentFile(upload_bytes), save=True)

    for error in preview.errors:
        MigrationRowError.objects.create(
            batch=batch,
            row_number=error.row_number,
            sheet_name=error.sheet_name,
            raw_payload=error.raw_payload,
            error_message=error.error_message,
        )

    for row in preview.valid_rows:
        if preview.import_type == MigrationBatch.ImportType.PRODUCTION_DASHBOARD_WORKBOOK:
            _persist_workbook_row(batch, row.row_number, row.normalized)
        else:
            _persist_row(batch, preview.import_type, row.row_number, row.normalized)
    return batch
